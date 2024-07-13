from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
excelfile = '德利佳沙田机房新规划.xlsx'
wb = load_workbook(excelfile)
ws = wb.active
