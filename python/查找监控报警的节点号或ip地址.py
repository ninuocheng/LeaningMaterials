import os.path
import time

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
excelfilename = '机器列表.xlsx'
if not os.path.exists(excelfilename):
    print(f'{excelfilename}不存在,请检查')
    time.sleep(5)
    exit()
wb = load_workbook(excelfilename)
ws = wb.active
a_lst = []
b_lst = []
for row in range(2,60):
    for col in range(1,14):
        char = get_column_letter(col)
        value = ws[char + str(row)].value
        a_lst.append(value)
    b_lst.append(a_lst)
    a_lst = []
def main():
    while True:
        menu()
        try:
            choice = int(input('请选择: '))
        except:
            print('输入有误,请重新选择')
            continue
        if choice == 0:
            exit()
        elif choice == 1:
            FindIpOrID()
        else:
            print('输入无效,请重新选择')
            continue
def menu():
    print('功能菜单'.center(30,'-'))
    print('0.退出，1.查找监控报警的ip或节点号: ')
def FindIpOrID():
    ipOrID = input('请输入要查找的IP或节点ID: ')
    Flag = False
    for item in b_lst:
        if ipOrID in item:
            Flag = True
            print(item)
    if not Flag:
        print(f'{ipOrID} not find')
if __name__ == '__main__':
    main()