import json
import time
import urllib.request
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
def main():
    while True:
        menu()
        try:
            choice = int(input('请选择: '))
        except:
            print('输入有误,请重新选择')
            continue
        if choice == 0:
            exit()
        elif choice == 1:
            MyAssets()
        else:
            print('输入有误,请重新选择')
            continue
def menu():
    print('JumpServer资产'.center(30,'='))
    print('功能菜单'.center(30,'-'))
    print('0.退出, 1.资产')
def MyAssets():
    while True:
        try:
            pagesNum = int(input('请输入页码数:'))
        except:
            print('输入有误,请重新输入')
            continue
        else:
            break
    excelfilename = 'asset.xlsx'
    wb = Workbook()
    ws = wb.active
    title = ['IP', '主机名', '远程端口']
    ws.append(title)
    for col in range(1,4):
        char = get_column_letter(col)
        ws[char + '1'].font = Font(bold=True,color='000000FF')
    for page in range(0,pagesNum):
        offset = page*15
        url = f'http://14.152.45.18:8078/api/v1/perms/users/assets/?offset={offset}&limit=15&display=1&draw=1'
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Cookie': 'Cookie: SESSION_COOKIE_NAME_PREFIX=jms_; X-JMS-ORG=00000000-0000-0000-0000-000000000002; jms_public_key="LS0tLS1CRUdJTiBQVUJMSUMgS0VZLS0tLS0KTUlHZk1BMEdDU3FHU0liM0RRRUJBUVVBQTRHTkFEQ0JpUUtCZ1FDdlFDMlBPUUhVQTBtdVBwQTlRK3Q1MG9lUApaL0U1cm1pZFRvSW9yVlROaDdBK0N2V2lUeHV4WDlJRzlxcllBekNrWHNQcVJlUzlUV3F5MVNWL21ocWxhVmxCCnZJSFg3SE9jb2Y0Ylc4N2EvSEJCdmRiVzEybU5ObzVOckt1NUNEVnFRODFZNFdFdW11N2RUY2tRaE9TZ2FiY1cKcDBNRzcwZ1liVDd2OGsySjJ3SURBUUFCCi0tLS0tRU5EIFBVQkxJQyBLRVktLS0tLQ=="; jms_session_expire=close; jms_csrftoken=U2Fx0XhMr9s57rtaQqD2QVaRocVSMMvYp9YpA4LPq7Qkx4okg1MKmHR0GQEhDgIP; jms_sessionid=6tkqebd5c6vb2l4gym2mdaf9nvj65pym',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'
        }
        request = urllib.request.Request(url=url,headers=headers)
        try:
            with urllib.request.urlopen(request) as response:
                content = json.load(response)
        except:
            print(f'请求的{url}访问失败,请检查')
            menu()
        count = content['count']
        results = content['results']
        for item in results:
            hostname = item['hostname']
            ip = item['ip']
            protocols = item['protocols']
            for i in protocols:
                port = int(i.split('/')[1])
                ws.append([hostname, ip, port])
                print(hostname, ip, port)
    try:
        wb.save(excelfilename)
    except PermissionError as error:
        print(error)
        menu()
    print(f'资产总计: {count},保存在{excelfilename}表格中')
if __name__ == '__main__':
        main()