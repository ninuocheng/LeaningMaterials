import json

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
filename = 'D:\\exel\机器列表.xlsx'
wb = load_workbook(filename)
ws = wb.active
# wsnames = wb.sheetnames
# print(wsnames)
# print(ws)
# print(ws.title)
lst = []
a_lst = []
for row in range(2,51):
    for col in range(1,14):
        char = get_column_letter(col)
        data = ws[char + str(row)].value
        lst.append(data)
    a_lst.append(lst)
    lst = []
print(a_lst,len(a_lst))
ipAddress = input('请输入要查询的ip: ')
for item in a_lst:
    if ipAddress == item[0]:
        print(item)
# b_lst = []
# for item in a_lst:
#     ip = item[0]
#     port = item[1]
#     minerID = item[2]
#     region = item[3]
#     user = item[4]
#     gpuNum = item[5]
#     lotusDaemon = item[9]
#     winning = item[10]
#     window = item[11]
#     sealing = item[12]
#     dct = {
#         'ip':ip,
#         'port':port,
#         'minerID':minerID,
#         'region':region,
#         'user':user,
#         'gpuNum':gpuNum,
#         'lotusDaemon':lotusDaemon,
#         'winning':winning,
#         'window':window,
#         'sealing':sealing
#     }
#     b_lst.append(dct)
# with open('abc',mode='a',encoding='utf-8') as afile:
#     afile.write(str(b_lst))